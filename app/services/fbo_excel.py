"""
ZELYONIY KORIDOR — Excel eksporti.

FBO ga tovar jo'natilganda, ba'zi jo'natmalar "yashil koridor" orqali
o'tadi. Buning uchun alohida Excel jadval to'ldiriladi.

Ustunlar (foydalanuvchi bergan namunaga aynan mos):
    Состав товара | Номер акта | Ссылка на акты | Кол-во SKU |
    Общее количество единиц | Сумма накладной |
    Дата отгрузки (по накладной) | Планируемая дата отгрузки

Muhim farq ikki sana orasida:
    "Дата отгрузки (по накладной)"   — Uzumning o'zidagi HAQIQIY sana
    "Планируемая дата отгрузки"      — FOYDALANUVCHI bergan sana

"Ссылка на акты" ustuni BO'SH qoldiriladi — bot havolani o'zi
yasay olmaydi (fayl qayerda saqlanishini bilmaydi), foydalanuvchi
buni qo'lda to'ldiradi.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

HEADERS = [
    "Состав товара",
    "Номер акта",
    "Ссылка на акты",
    "Кол-во SKU",
    "Общее количество единиц",
    "Сумма накладной",
    "Дата отгрузки (по накладной)",
    "Планируемая дата отгрузки",
]


def _fmt_date(value: Any) -> str:
    """Uzum sanasini DD.MM.YYYY ko'rinishiga o'giradi."""
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    text = str(value)
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:19], fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    return text[:10]


def build(
    invoice: dict[str, Any],
    products: list[dict[str, Any]],
    planned_date: str,
) -> bytes:
    """Bitta FBO yuk xati uchun Excel — orqaga moslik uchun saqlanadi."""
    return build_many([invoice], planned_date)


def build_many(invoices: list[dict[str, Any]], planned_date: str) -> bytes:
    """
    BARCHA (odatda yangi/qabul qilinmagan) FBO yuk xatlari — BITTA
    Excel faylida, har biri alohida qatorda.

    "Планируемая дата отгрузки" — hammasiga BIR XIL, foydalanuvchi
    bergan sana. Har aktning o'z "Дата отгрузки (по накладной)" esa
    Uzumdan alohida-alohida olinadi.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Zelyoniy koridor"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="7B2D8E")
    body_font = Font(name="Arial")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for row_i, invoice in enumerate(invoices, start=2):
        products = invoice.get("products") or []
        composition = "; ".join(
            f"{p['name']} x{p['to_stock']}" for p in products
        ) if products else ""
        sku_count = len(products) if products else 0
        total_units = (
            sum(p["to_stock"] for p in products)
            if products else invoice.get("total_to_stock", 0)
        )

        row = [
            composition,
            invoice.get("number"),
            "",  # Ссылка на акты — qo'lda to'ldiriladi
            sku_count,
            total_units,
            invoice.get("total_price", 0),
            _fmt_date(invoice.get("date_created")),
            planned_date,
        ]
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.font = body_font
            cell.alignment = wrap

    widths = [45, 14, 22, 12, 16, 16, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
